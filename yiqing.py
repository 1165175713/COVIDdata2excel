# -*- coding: utf-8 -*-
"""
Created on Fri May 15 09:26:38 2020

@author: fyswi
"""

import requests
from bs4 import BeautifulSoup 
import openpyxl


class COVIDDATA:
    content = ""
    table = ""
    def __init__(self):
        self.url = url
    
    def htmlfile(self):
        r = requests.get(self.url)
        self.content=r.content
        return self.content

    def htmltable(self):
        soup = BeautifulSoup(self.htmlfile())
        
        return self.table
    
    def Head(self):
        self.table = self.htmltable()
        tablehead = self.table.thead.find_all("th")
    
        head = []
        for i in tablehead:
            head.append(i.get_text())
        del head[2]
        del head[2]

        return head[2:]
    
    

def req(url):
    r = requests.get(url)
    with open("yiqing.html","wb") as f:
        f.write(r.content)


def Head(table):
    tablehead = table.thead.find_all("th")
    
    head = []
    for i in tablehead:
        head.append(i.get_text())
    del head[2]
    del head[2]
    return head[2:]

def Body(table):
    body = []
    tablebody = table.tbody.find_all("tr")
    for tr in tablebody:
        tmp = []
        trs = tr.find_all('td')
        del trs[3]
        del trs[3]
        for td in trs:
            tmp.append(td.get_text())
        body.append(tmp[1:])
    return body

def Body_China(table):
    body = []
    tablebody = table.tbody.find_all("tr")
    for tr in tablebody:
        tmp = []
        trs = tr.find_all('td')
        del trs[3]
        del trs[3]
        for td in trs:
            tmp.append(td.get_text())
        if tmp[2] == "China":
            body.append(tmp[1:])
    return body

def Body_country(body):
    tmp = []
    res = []
    length = len(body[0])-2
    countries = {'United Kingdom':[0]*length, 'Denmark':[0]*length, 'China':[0]*length, 'Australia':[0]*length, 'France':[0]*length, 'Netherlands':[0]*length, 'Canada':[0]*length}
    
    for i in body:
        if i[1] in countries.keys():
            tmp.append(i)
        else:
            res.append(i)
    
    for i in tmp:
        
        for j in range(length):
            countries[i[1]][j] = countries[i[1]][j] + int(i[j+2])   
             
    for i,j in countries.items():
        a = ["",i]
        a.extend(j)
        res.append(a)
    return res

if __name__ == "__main__":  
    url1="https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_confirmed_global.csv"
    url2="https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_deaths_global.csv"   
    url3="https://github.com/CSSEGISandData/COVID-19/blob/master/csse_covid_19_data/csse_covid_19_time_series/time_series_covid19_recovered_global.csv"
    url = url3
    req(url)
    soup = BeautifulSoup(open('yiqing.html','rb'))
    table = soup.table
    head = Head(table)
    body = Body(table)

    #body = Body_China(table)
    #body = Body_country(body) 

    # tmp = []
    # for i in body:
    #     tmp.append(i[1])
    # import pickle
    # with open("tmp2","wb") as f:
    #     pickle.dump(tmp, f)

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0) 
    for i in range(len(body)):
        for j in range(len(head)):
            outws.cell(i*len(head)+j+1, 1).value = body[i][0] 
            outws.cell(i*len(head)+j+1, 2).value = body[i][1] 
            outws.cell(i*len(head)+j+1, 3).value = head[j]
            if j==0:
                outws.cell(i*len(head)+j+1, 4).value = body[i][j+2]
            else:
                outws.cell(i*len(head)+j+1, 4).value = int(body[i][j+2])-int(body[i][j+1])
            try:
                outws.cell(i*len(head)+j+1, 5).value = body[i][j+2]
            except:
                print(i,j)
         
    outwb.save("yiqing"+url[111:]+".xlsx")
