# -*- coding: utf-8 -*-
"""
Created on Thu May 21 11:28:28 2020

@author: fyswi
"""




from pandas import *
from openpyxl import *


def addallsame(csvdata):
    res = {}
    for i in range(csvdata.shape[0]):
        if csvdata.iloc[i][0] not in res.keys():
            res[csvdata.iloc[i][0]] =  csvdata.iloc[i][1:]
        else:
            res[csvdata.iloc[i][0]] =  res[csvdata.iloc[i][0]]+csvdata.iloc[i][1:]
    return res

if __name__=="__main__":
    csv1 = read_csv("time_series_covid19_confirmed_US.csv")
    csv2 = read_csv("time_series_covid19_deaths_US.csv")
    
    confirmed = addallsame(csv1)
    deaths = addallsame(csv2)
    wb = Workbook()
    ws = wb.active
    
    
    import datetime
    counter = -1
    for j in confirmed.keys():
        counter += 1
        a = datetime.date(2020, 1, 22)
        for k in range(len(confirmed[j])): 
            ws.cell(counter*len(confirmed["Alabama"])+k+1,1).value = j
            ws.cell(counter*len(confirmed["Alabama"])+k+1,2).value = a.strftime("%m/%d/%Y")
            if k==0:
                ws.cell(counter*len(confirmed["Alabama"])+k+1,3).value = confirmed[j][k]
            else:
                ws.cell(counter*len(confirmed["Alabama"])+k+1,3).value = confirmed[j][k] - confirmed[j][k-1]
            ws.cell(counter*len(confirmed["Alabama"])+k+1,4).value = confirmed[j][k]
            ws.cell(counter*len(confirmed["Alabama"])+k+1,5).value = deaths[j][k]
            a = a+datetime.timedelta(days=1)
    wb.save("us1.xlsx")       
        

