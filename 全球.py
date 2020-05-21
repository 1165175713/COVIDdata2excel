from openpyxl import *

def findd(xlsxsheet,tmp1,tmp2,tmp3):
    for i in range(xlsxsheet.max_row):
        if xlsxsheet.cell(i+1,1).value == tmp1 and xlsxsheet.cell(i+1,2).value == tmp2 and xlsxsheet.cell(i+1,3).value == tmp3:
            return xlsxsheet.cell(i+1,4).value,xlsxsheet.cell(i+1,5).value
    return ""


if __name__ == "__main__":  

    xlsx1 = load_workbook('yiqing_covid19_confirmed_global.csv.xlsx')
    xlsx2 = load_workbook('yiqing_covid19_deaths_global.csv.xlsx')
    xlsx3 = load_workbook('yiqing_covid19_recovered_global.csv.xlsx')

    xlsx1sheet = xlsx1.get_sheet_by_name("Sheet1")
    xlsx2sheet = xlsx2.get_sheet_by_name("Sheet1")
    xlsx3sheet = xlsx3.get_sheet_by_name("Sheet1")

    outwb = Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0) 
    for i in range(xlsx3sheet.max_row):
        print(i)
        outws.cell(i+1, 1).value = xlsx3sheet.cell(i+1,1).value
        outws.cell(i+1, 2).value = xlsx3sheet.cell(i+1,2).value
        outws.cell(i+1, 3).value = xlsx3sheet.cell(i+1,3).value
        tmp = findd(xlsx1sheet,xlsx3sheet.cell(i+1,1).value,xlsx3sheet.cell(i+1,2).value,xlsx3sheet.cell(i+1,3).value)
        if tmp!="":
            outws.cell(i+1, 4).value = tmp[0]
            outws.cell(i+1, 5).value = tmp[1]
        
        tmp = findd(xlsx2sheet,xlsx3sheet.cell(i+1,1).value,xlsx3sheet.cell(i+1,2).value,xlsx3sheet.cell(i+1,3).value)
        if tmp!="":
            outws.cell(i+1, 6).value = tmp[0]
            outws.cell(i+1, 7).value = tmp[1]

        outws.cell(i+1, 8).value = xlsx3sheet.cell(i+1,1).value
    outwb.save("1111.xlsx")
            



    